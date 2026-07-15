import copy
import io
import json
import sys
import tempfile
import unittest
from pathlib import Path
from unittest import mock


ROOT = Path(__file__).resolve().parents[1]
SCRIPTS = ROOT / "scripts"
sys.path.insert(0, str(SCRIPTS))

import video_timeline
from test_video_analysis_pipeline import valid_analysis
import write_feishu_sheet


class FakeRunner:
    def __init__(self, fail_operation=None, existing_title=None, cleanup_fails=False):
        self.calls = []
        self.fail_operation = fail_operation
        self.existing_title = existing_title
        self.cleanup_fails = cleanup_fails
        self.info_count = 0

    def run(self, operation, args, *, output_path=None):
        self.calls.append((operation, list(args)))
        if operation == "delete" and self.cleanup_fails:
            raise write_feishu_sheet.LarkCliError("cleanup failed")
        if (
            operation == self.fail_operation
            or (self.fail_operation == "post_create_metadata" and operation == "metadata" and self.info_count == 1)
            or (self.fail_operation == "final_metadata" and operation == "metadata" and self.info_count == 2)
        ):
            raise write_feishu_sheet.LarkCliError(f"{operation} failed")
        if operation == "metadata":
            self.info_count += 1
            merges = []
            for name, command in self.calls:
                if name == "merge":
                    payload = json.loads(command[command.index("--data") + 1])
                    merges.append(payload["range"])
            sheets = [{"sheet_id": "old", "title": "原页签", "merges": []}]
            if self.existing_title:
                sheets.append({"sheet_id": "conflict", "title": self.existing_title})
            if self.info_count > 1:
                sheets.append(
                    {
                        "sheet_id": "new-sheet",
                        "title": "前期体验拆解",
                        "merges": merges,
                    }
                )
            return {"spreadsheet_token": "sht-test", "sheets": sheets}
        if operation == "create":
            return {"data": {"replies": [{"addSheet": {"properties": {"sheetId": "new-sheet"}}}]}}
        if operation == "read":
            values = []
            for name, command in self.calls:
                if name == "write":
                    if "--values" in command:
                        values.extend(json.loads(command[command.index("--values") + 1]))
                    else:
                        cells = json.loads(command[command.index("--cells-json") + 1])
                        values.extend(
                            [
                                [
                                    cell.get("value") if isinstance(cell, dict) else None
                                    for cell in row
                                ]
                                for row in cells
                            ]
                        )
            return {"values": values}
        if operation == "export":
            Path(output_path).write_bytes(b"fake-xlsx")
            return {"file_path": str(output_path)}
        return {"code": 0}


def operation_names(runner):
    return [name for name, _ in runner.calls]


def make_export_workbook(path, sheet_title, layout):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_title
    for row in layout["matrix"]:
        worksheet.append(row)
    for cell_range in layout["merges"]:
        worksheet.merge_cells(cell_range)
    for row, pixels in layout["row_heights"].items():
        worksheet.row_dimensions[row].height = pixels * (2 / 3)
    for column, pixels in layout["column_widths"].items():
        worksheet.column_dimensions[
            write_feishu_sheet.column_letter(column)
        ].width = write_feishu_sheet.excel_width_for_pixels(pixels)
    worksheet["A1"].fill = PatternFill(
        fill_type="solid", fgColor=write_feishu_sheet.MAIN_YELLOW[1:]
    )
    worksheet["A1"].font = Font(
        size=36, bold=True, color=write_feishu_sheet.TITLE_FORE[1:]
    )
    workbook.save(path)
    workbook.close()


class SheetLayoutTests(unittest.TestCase):
    def test_feishu_export_column_width_uses_eight_pixels_per_excel_unit(self):
        self.assertAlmostEqual(27.5, write_feishu_sheet.excel_width_for_pixels(220))
        self.assertAlmostEqual(17.5, write_feishu_sheet.excel_width_for_pixels(140))

    def test_two_hours_seventeen_minutes_has_44_timeline_columns(self):
        analysis = valid_analysis((2 * 60 + 17) * 60)
        layout = write_feishu_sheet.build_sheet_layout(analysis)
        self.assertEqual(44, layout["timeline_column_count"])
        self.assertEqual(44, len(analysis["slices"]))
        self.assertEqual(45, layout["column_count"])

    def test_layout_contains_all_seven_dimension_rows(self):
        layout = write_feishu_sheet.build_sheet_layout(valid_analysis())
        self.assertEqual(
            ["阶段目标", "任务链", "核心循环", "渐进体验", "地图体验", "经济体验", "剧情轴"],
            list(layout["rows"]),
        )
        self.assertEqual(
            [
                "阶段目标 (Experience Goal)",
                "任务链 (Quest Chain)",
                "核心循环 (Core Loop)",
                "渐进体验预期 (New Content)",
                "地图体验预期 (Map Progress)",
                "经济体验预期 (Eco Progress)",
                "剧情轴",
            ],
            [row[0] for row in layout["matrix"][2:]],
        )
        self.assertTrue(set(layout["rows"].values()).issubset(layout["row_heights"]))
        self.assertEqual("36pt/1.5", layout["styles"][0]["style"]["font"]["fontSize"])
        self.assertEqual("11pt/1.5", layout["base_style"]["style"]["font"]["fontSize"])

    def test_consecutive_stages_create_dynamic_merges(self):
        analysis = valid_analysis(180.0)
        for index, item in enumerate(analysis["slices"]):
            if index < 2:
                item["dimensions"]["阶段目标"] = {
                    "fact": "出现移动教学",
                    "inference": "目标是学会移动",
                }
                item["stage_range"] = {
                    "stage_id": "tutorial",
                    "name": "教学",
                    "start": 0.0,
                    "end": 120.0,
                }
            else:
                item["dimensions"]["阶段目标"] = {
                    "fact": "进入战斗",
                    "inference": "目标转为战斗",
                }
                item["stage_range"] = {
                    "stage_id": "combat",
                    "name": "战斗",
                    "start": 120.0,
                    "end": 180.0,
                }
        layout = write_feishu_sheet.build_sheet_layout(analysis)
        stage_row = layout["rows"]["阶段目标"]
        self.assertIn(f"B{stage_row}:C{stage_row}", layout["merges"])
        self.assertNotIn(f"B{stage_row}:D{stage_row}", layout["merges"])

    def test_cells_only_show_concise_fact_and_omit_internal_analysis_markers(self):
        analysis = valid_analysis()
        item = analysis["slices"][0]
        item["narrative_climax"] = {"judgement": "climax", "reason": "首领登场"}
        item["flow"] = {"judgement": "flow_peak", "reason": "挑战匹配"}
        layout = write_feishu_sheet.build_sheet_layout(analysis)
        all_cells = "\n".join(str(cell or "") for row in layout["matrix"] for cell in row)
        self.assertNotIn("事实：", all_cells)
        self.assertNotIn("推断：", all_cells)
        self.assertNotIn("高潮/低谷：", all_cells)
        self.assertNotIn("心流：", all_cells)
        narrative = layout["matrix"][layout["rows"]["剧情轴"] - 1][1]
        self.assertEqual("冲突被引入", narrative)

    def test_empty_dimension_fact_produces_empty_feishu_cell(self):
        analysis = valid_analysis()
        analysis["slices"][0]["dimensions"]["经济体验"] = {
            "fact": "",
            "inference": "",
        }
        layout = write_feishu_sheet.build_sheet_layout(analysis)
        value = layout["matrix"][layout["rows"]["经济体验"] - 1][1]
        self.assertEqual("", value)

    def test_styles_and_heights_only_reference_final_layout(self):
        layout = write_feishu_sheet.build_sheet_layout(valid_analysis(120.0))
        max_row = len(layout["matrix"])
        max_col = layout["column_count"]
        for row in layout["row_heights"]:
            self.assertLessEqual(row, max_row)
        for style in layout["styles"]:
            bounds = write_feishu_sheet.a1_bounds(style["range"])
            self.assertLessEqual(bounds[2], max_row)
            self.assertLessEqual(bounds[3], max_col)

    def test_long_text_increases_dynamic_row_height(self):
        analysis = valid_analysis()
        analysis["slices"][0]["dimensions"]["任务链"]["fact"] = "很长的文本" * 80
        layout = write_feishu_sheet.build_sheet_layout(analysis)
        self.assertGreater(layout["row_heights"][layout["rows"]["任务链"]], 36)


class WriteWorkflowTests(unittest.TestCase):
    def test_real_info_response_nested_sheets_and_grid_properties_are_normalized(self):
        metadata = {
            "data": {
                "sheets": {
                    "sheets": [
                        {
                            "sheet_id": "nested-sheet",
                            "title": "寒霜拆解",
                            "grid_properties": {
                                "row_count": 200,
                                "column_count": 20,
                            },
                        }
                    ]
                },
                "spreadsheet": {"spreadsheet": {"token": "sht-nested"}},
            }
        }
        sheets = write_feishu_sheet._sheets(metadata)
        self.assertEqual("nested-sheet", sheets[0]["sheet_id"])
        self.assertEqual((200, 20), write_feishu_sheet._sheet_grid_size(sheets[0]))
        self.assertEqual(
            "sht-nested",
            write_feishu_sheet._spreadsheet_token("https://example/sheets/fallback", metadata),
        )

    def test_real_info_merge_indexes_are_converted_to_a1_ranges(self):
        sheet = {
            "merges": [
                {
                    "start_row_index": 0,
                    "end_row_index": 0,
                    "start_column_index": 0,
                    "end_column_index": 2,
                }
            ]
        }
        self.assertEqual({"A1:C1"}, write_feishu_sheet._extract_merges(sheet))

    def test_deprecated_read_response_value_range_is_normalized(self):
        payload = {
            "data": {
                "valueRange": {
                    "range": "sheet!A1:B2",
                    "values": [["标题", None], ["阶段目标", "内容"]],
                }
            }
        }
        self.assertEqual(
            [["标题", None], ["阶段目标", "内容"]],
            write_feishu_sheet._read_values(payload),
        )

    def test_preflight_is_read_only_and_rejects_existing_title(self):
        runner = FakeRunner()
        result = write_feishu_sheet.write_analysis(
            valid_analysis(),
            "sht-test",
            "新的拆解页签",
            runner=runner,
            preflight=True,
        )
        self.assertTrue(result["preflight"])
        self.assertTrue(result["title_available"])
        self.assertEqual(["metadata"], operation_names(runner))

        conflict_runner = FakeRunner(existing_title="前期体验拆解")
        with self.assertRaisesRegex(
            write_feishu_sheet.SheetTitleConflict, "新的唯一页签名|取消"
        ):
            write_feishu_sheet.write_analysis(
                valid_analysis(),
                "sht-test",
                "前期体验拆解",
                runner=conflict_runner,
                preflight=True,
            )
        self.assertEqual(["metadata"], operation_names(conflict_runner))

    def test_same_title_conflict_does_not_write(self):
        runner = FakeRunner(existing_title="前期体验拆解")
        with self.assertRaisesRegex(
            write_feishu_sheet.SheetTitleConflict, "新的唯一页签名|取消"
        ):
            write_feishu_sheet.write_analysis(
                valid_analysis(), "sht-test", "前期体验拆解", runner=runner
            )
        self.assertEqual(["metadata"], operation_names(runner))

    def test_dry_run_returns_complete_plan_without_runner_call(self):
        runner = FakeRunner()
        result = write_feishu_sheet.write_analysis(
            valid_analysis(),
            "sht-test",
            "前期体验拆解",
            runner=runner,
            dry_run=True,
            batch_size=1,
        )
        self.assertEqual([], runner.calls)
        self.assertTrue(result["dry_run"])
        self.assertEqual(
            ["metadata", "create", "expand", "write", "merge", "column_width", "base_style",
             "local_style", "row_height", "read", "metadata", "export"],
            result["operation_order"],
        )
        self.assertTrue(result["operations"])

    def test_full_call_order_and_batch_limit(self):
        runner = FakeRunner()
        with mock.patch(
            "write_feishu_sheet.verify_export_xlsx",
            return_value={"available": True, "passed": True},
        ):
            result = write_feishu_sheet.write_analysis(
                valid_analysis(120.0),
                "https://example.feishu.cn/sheets/sht-test",
                "前期体验拆解",
                runner=runner,
                batch_size=1,
            )
        names = operation_names(runner)
        self.assertEqual(["metadata", "create", "metadata"], names[:3])
        self.assertNotIn("expand", names)
        self.assertEqual(9, names.count("write"))
        self.assertLess(names.index("write"), names.index("merge"))
        self.assertLess(names.index("merge"), names.index("column_width"))
        self.assertLess(names.index("column_width"), names.index("base_style"))
        self.assertLess(names.index("base_style"), names.index("local_style"))
        self.assertLess(names.index("local_style"), names.index("row_height"))
        self.assertLess(names.index("row_height"), names.index("read"))
        self.assertLess(names.index("read"), names.index("export"))
        self.assertTrue(result["verification"]["passed"])

    def test_write_uses_cells_set_file_payload_instead_of_deprecated_values(self):
        runner = FakeRunner(fail_operation="write")
        with self.assertRaises(write_feishu_sheet.WriteTransactionError):
            write_feishu_sheet.write_analysis(
                valid_analysis(),
                "sht-test",
                "前期体验拆解",
                runner=runner,
                batch_size=1,
            )

        write_args = next(args for operation, args in runner.calls if operation == "write")
        self.assertIn("+cells-set", write_args)
        self.assertIn("--cells-json", write_args)
        self.assertNotIn("+write", write_args)
        self.assertNotIn("--values", write_args)

    def test_create_payload_and_insert_dimension_follow_real_contract(self):
        runner = FakeRunner()
        with mock.patch(
            "write_feishu_sheet.verify_export_xlsx",
            return_value={"available": True, "passed": True},
        ):
            write_feishu_sheet.write_analysis(
                valid_analysis((2 * 60 + 17) * 60),
                "sht-test",
                "前期体验拆解",
                runner=runner,
            )
        create_args = next(args for name, args in runner.calls if name == "create")
        create_payload = json.loads(create_args[create_args.index("--data") + 1])
        properties = create_payload["requests"][0]["addSheet"]["properties"]
        self.assertEqual({"title": "前期体验拆解"}, properties)
        expand_args = next(args for name, args in runner.calls if name == "expand")
        expand_payload = json.loads(expand_args[expand_args.index("--data") + 1])
        self.assertTrue(
            any(value.endswith("/insert_dimension_range") for value in expand_args)
        )
        dimension = expand_payload["dimension"]
        self.assertEqual("COLUMNS", dimension["majorDimension"])
        self.assertEqual(0, dimension["startIndex"])
        self.assertEqual(20, dimension["endIndex"])
        requests = write_feishu_sheet._dimension_insert_requests(
            "new-sheet", 120, 45, current_rows=100, current_columns=20
        )
        self.assertEqual(["COLUMNS", "COLUMNS", "ROWS"], [
            request["dimension"]["majorDimension"] for request in requests
        ])
        current_sizes = {"COLUMNS": 20, "ROWS": 100}
        inserted = {"COLUMNS": 0, "ROWS": 0}
        for request in requests:
            request_dimension = request["dimension"]
            major = request_dimension["majorDimension"]
            self.assertLessEqual(request_dimension["endIndex"], current_sizes[major])
            count = request_dimension["endIndex"] - request_dimension["startIndex"]
            inserted[major] += count
            current_sizes[major] += count
        self.assertEqual({"COLUMNS": 25, "ROWS": 20}, inserted)

    def test_each_post_create_failure_deletes_new_sheet_and_returns_two(self):
        stages = [
            "post_create_metadata", "expand", "write", "merge", "column_width",
            "base_style", "local_style", "row_height", "read", "final_metadata", "export",
        ]
        with tempfile.TemporaryDirectory() as directory:
            source = Path(directory) / "analysis.json"
            source.write_text(json.dumps(valid_analysis(), ensure_ascii=False), "utf-8")
            for stage in stages:
                with self.subTest(stage=stage):
                    stderr = io.StringIO()
                    runner = FakeRunner(fail_operation=stage)
                    with mock.patch("sys.stderr", stderr):
                        code = write_feishu_sheet.main(
                            [str(source), "sht-test", "前期体验拆解"], runner=runner
                        )
                    self.assertEqual(2, code)
                    self.assertIn("error", json.loads(stderr.getvalue()))
                    self.assertEqual("delete", operation_names(runner)[-1])

    def test_cleanup_failure_is_explicit_in_error_json(self):
        with tempfile.TemporaryDirectory() as directory:
            source = Path(directory) / "analysis.json"
            source.write_text(json.dumps(valid_analysis(), ensure_ascii=False), "utf-8")
            stderr = io.StringIO()
            runner = FakeRunner(fail_operation="write", cleanup_fails=True)
            with mock.patch("sys.stderr", stderr):
                code = write_feishu_sheet.main(
                    [str(source), "sht-test", "前期体验拆解"], runner=runner
                )
        payload = json.loads(stderr.getvalue())
        self.assertEqual(2, code)
        self.assertEqual("failed", payload["error"]["cleanup"]["status"])
        self.assertIn("cleanup failed", payload["error"]["cleanup"]["message"])

    def test_cli_runner_rejects_non_json_output_with_json_protocol_error(self):
        completed = mock.Mock(returncode=0, stdout="not json", stderr="")
        with mock.patch("write_feishu_sheet.subprocess.run", return_value=completed) as run:
            with self.assertRaises(write_feishu_sheet.LarkCliError):
                write_feishu_sheet.LarkRunner(timeout=1).run("metadata", ["sheets", "+info"])
        command = run.call_args.args[0]
        self.assertIsInstance(command, list)
        self.assertEqual("utf-8", run.call_args.kwargs["encoding"])
        self.assertEqual("replace", run.call_args.kwargs["errors"])
        self.assertEqual(1, run.call_args.kwargs["timeout"])

    def test_cli_runner_materializes_unicode_cells_json_as_file_argument(self):
        observed = {}

        def invoke(command, **kwargs):
            observed["command"] = command
            if "--cells" in command:
                reference = command[command.index("--cells") + 1]
                observed["reference"] = reference
                observed["payload"] = json.loads(
                    Path(reference.removeprefix("@")).read_text(encoding="utf-8")
                )
            return mock.Mock(returncode=0, stdout='{"ok": true}', stderr="")

        cells = [[{"value": "三冰拆解"}, {"value": ""}]]
        with mock.patch("write_feishu_sheet.subprocess.run", side_effect=invoke):
            write_feishu_sheet.LarkRunner(timeout=1).run(
                "write",
                [
                    "sheets",
                    "+cells-set",
                    "--cells-json",
                    json.dumps(cells, ensure_ascii=False),
                ],
            )

        self.assertIn("--cells", observed["command"])
        self.assertNotIn("--cells-json", observed["command"])
        self.assertTrue(observed["reference"].startswith("@"))
        self.assertFalse(Path(observed["reference"][1:]).is_absolute())
        self.assertEqual(cells, observed["payload"])
        self.assertFalse(Path(observed["reference"][1:]).exists())

    def test_invalid_analysis_fails_before_network(self):
        runner = FakeRunner()
        analysis = valid_analysis()
        del analysis["slices"][0]["dimensions"]["经济体验"]
        with self.assertRaises(Exception):
            write_feishu_sheet.write_analysis(
                analysis, "sht-test", "前期体验拆解", runner=runner
            )
        self.assertEqual([], runner.calls)

    def test_export_file_is_always_removed(self):
        runner = FakeRunner()
        observed = []

        def inspect(path, *args, **kwargs):
            observed.append(Path(path))
            self.assertTrue(Path(path).exists())
            return {"available": True, "passed": True}

        with mock.patch("write_feishu_sheet.verify_export_xlsx", side_effect=inspect):
            write_feishu_sheet.write_analysis(
                valid_analysis(), "sht-test", "前期体验拆解", runner=runner
            )
        self.assertEqual(1, len(observed))
        self.assertFalse(observed[0].exists())
        export_args = next(args for name, args in runner.calls if name == "export")
        cli_output = export_args[export_args.index("--output-path") + 1]
        self.assertFalse(Path(cli_output).is_absolute())

    def test_style_verification_failure_returns_two_and_cleans_export(self):
        runner = FakeRunner()
        observed = []

        def fail_verification(path, *args, **kwargs):
            observed.append(Path(path))
            raise ValueError("bad workbook style")

        with tempfile.TemporaryDirectory() as directory:
            source = Path(directory) / "analysis.json"
            source.write_text(json.dumps(valid_analysis(), ensure_ascii=False), "utf-8")
            stderr = io.StringIO()
            with (
                mock.patch(
                    "write_feishu_sheet.verify_export_xlsx",
                    side_effect=fail_verification,
                ),
                mock.patch("sys.stderr", stderr),
            ):
                code = write_feishu_sheet.main(
                    [str(source), "sht-test", "前期体验拆解"], runner=runner
                )
        self.assertEqual(2, code)
        self.assertIn("样式验证失败", json.loads(stderr.getvalue())["error"]["message"])
        self.assertEqual(1, len(observed))
        self.assertFalse(observed[0].exists())

    def test_missing_openpyxl_is_a_verification_failure(self):
        with (
            tempfile.TemporaryDirectory() as directory,
            mock.patch.dict(sys.modules, {"openpyxl": None}),
        ):
            path = Path(directory) / "export.xlsx"
            path.write_bytes(b"xlsx")
            with self.assertRaisesRegex(
                write_feishu_sheet.VerificationError, "openpyxl"
            ):
                write_feishu_sheet.verify_export_xlsx(
                    path, "前期体验拆解", write_feishu_sheet.build_sheet_layout(valid_analysis())
                )

    def test_export_verification_checks_all_required_visual_properties(self):
        analysis = valid_analysis()
        analysis["slices"][0]["dimensions"]["任务链"]["fact"] = "长文本" * 120
        layout = write_feishu_sheet.build_sheet_layout(analysis)
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "export.xlsx"
            make_export_workbook(path, "前期体验拆解", layout)
            result = write_feishu_sheet.verify_export_xlsx(
                path, "前期体验拆解", layout
            )
        self.assertTrue(result["passed"])
        self.assertEqual(
            {
                "key_merges",
                "title_timeline_dimension_row_heights",
                "long_text_row_heights",
                "title_fill",
                "title_font_color",
                "title_font_size",
                "timeline_and_dimension_column_widths",
            },
            set(result["checks"]),
        )

    def test_export_verification_accepts_small_positive_row_height_drift(self):
        from openpyxl import load_workbook

        layout = write_feishu_sheet.build_sheet_layout(valid_analysis())
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "export.xlsx"
            make_export_workbook(path, "前期体验拆解", layout)
            workbook = load_workbook(path)
            worksheet = workbook["前期体验拆解"]
            worksheet.row_dimensions[1].height = layout["row_heights"][1] * (2 / 3) + 5.5
            workbook.save(path)
            workbook.close()
            result = write_feishu_sheet.verify_export_xlsx(
                path, "前期体验拆解", layout
            )

        self.assertTrue(result["passed"])

    def test_export_verification_rejects_each_visual_mismatch(self):
        from openpyxl import load_workbook

        analysis = valid_analysis()
        analysis["slices"][0]["dimensions"]["任务链"]["fact"] = "长文本" * 120
        layout = write_feishu_sheet.build_sheet_layout(analysis)

        def set_wrong_title_size(worksheet):
            font = copy.copy(worksheet["A1"].font)
            font.sz = 12
            worksheet["A1"].font = font

        mutations = {
            "关键合并": lambda ws: ws.unmerge_cells(layout["merges"][0]),
            "行高": lambda ws: setattr(ws.row_dimensions[1], "height", 5),
            "标题底色": lambda ws: setattr(ws["A1"].fill.fgColor, "rgb", "FFFFFFFF"),
            "标题字色": lambda ws: setattr(ws["A1"].font.color, "rgb", "FFFFFFFF"),
            "标题字号": set_wrong_title_size,
            "时间轴与维度列宽": lambda ws: setattr(ws.column_dimensions["B"], "width", 5),
            "长文本行高": lambda ws: setattr(
                ws.row_dimensions[layout["rows"]["任务链"]], "height", 12
            ),
        }
        with tempfile.TemporaryDirectory() as directory:
            for label, mutate in mutations.items():
                with self.subTest(label=label):
                    path = Path(directory) / f"{label}.xlsx"
                    make_export_workbook(path, "前期体验拆解", layout)
                    workbook = load_workbook(path)
                    worksheet = workbook["前期体验拆解"]
                    mutate(worksheet)
                    workbook.save(path)
                    workbook.close()
                    with self.assertRaisesRegex(
                        write_feishu_sheet.VerificationError, label
                    ):
                        write_feishu_sheet.verify_export_xlsx(
                            path, "前期体验拆解", layout
                        )


if __name__ == "__main__":
    unittest.main()
