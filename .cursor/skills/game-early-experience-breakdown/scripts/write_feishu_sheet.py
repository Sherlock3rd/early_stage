#!/usr/bin/env python3
"""Write a validated early-game analysis to a new Feishu Sheet tab."""

from __future__ import annotations

import argparse
import json
import math
import os
import re
import subprocess
import sys
import tempfile
from pathlib import Path
from typing import Any

try:
    from .analysis_model import DIMENSION_KEYS, validate_analysis
    from .runtime_utils import emit_json_error
except ImportError:
    from analysis_model import DIMENSION_KEYS, validate_analysis
    from runtime_utils import emit_json_error


DEFAULT_BATCH_SIZE = 20
MAX_BATCH_SIZE = 100
MAX_JSON_ARGUMENT_CHARS = 24_000
DEFAULT_TIMEOUT = 45.0
DEFAULT_SHEET_COLUMNS = 20
DEFAULT_SHEET_ROWS = 100
TITLE = "游戏前期体验拆解"
MAIN_YELLOW = "#FAC603"
LIGHT_YELLOW = "#FAF1D1"
DARK_GREY = "#3A3E43"
WHITE = "#FFFFFF"
TITLE_FORE = "#000000"
NARRATIVE_LABELS = {"none": "无", "climax": "高潮", "low": "低谷"}
FLOW_LABELS = {"none": "无", "flow_peak": "心流高点"}
DIMENSION_LABELS = {
    "阶段目标": "阶段目标 (Experience Goal)",
    "任务链": "任务链 (Quest Chain)",
    "核心循环": "核心循环 (Core Loop)",
    "渐进体验": "渐进体验预期 (New Content)",
    "地图体验": "地图体验预期 (Map Progress)",
    "经济体验": "经济体验预期 (Eco Progress)",
    "剧情轴": "剧情轴",
}


class LarkCliError(RuntimeError):
    """Raised when lark-cli fails or emits an invalid response."""


class SheetTitleConflict(RuntimeError):
    """Raised when the requested new tab title already exists."""


class VerificationError(RuntimeError):
    """Raised when post-write content or formatting verification fails."""


class WriteTransactionError(RuntimeError):
    """Wrap a post-create failure and the best-effort cleanup outcome."""

    def __init__(self, cause: BaseException, cleanup_error: BaseException | None = None):
        self.cause = cause
        self.cleanup_error = cleanup_error
        message = f"{type(cause).__name__}: {cause}"
        if cleanup_error is not None:
            message += f"；清理新页签失败: {type(cleanup_error).__name__}: {cleanup_error}"
        super().__init__(message)


def column_letter(index: int) -> str:
    if index < 1:
        raise ValueError("列索引必须大于 0")
    value = ""
    while index:
        index, remainder = divmod(index - 1, 26)
        value = chr(65 + remainder) + value
    return value


def column_index(letters: str) -> int:
    result = 0
    for character in letters.upper():
        result = result * 26 + ord(character) - 64
    return result


def excel_width_for_pixels(pixels: int | float) -> float:
    """Approximate Feishu-exported Excel width for a fixed pixel width."""
    return max(0.0, float(pixels) / 8.0)


def a1_bounds(cell_range: str) -> tuple[int, int, int, int]:
    raw = cell_range.split("!", 1)[-1]
    match = re.fullmatch(r"([A-Z]+)(\d+)(?::([A-Z]+)(\d+))?", raw, re.I)
    if not match:
        raise ValueError(f"非法 A1 范围: {cell_range}")
    start_col, start_row, end_col, end_row = match.groups()
    return (
        int(start_row),
        column_index(start_col),
        int(end_row or start_row),
        column_index(end_col or start_col),
    )


def _format_seconds(value: float) -> str:
    seconds = int(round(value))
    hours, remainder = divmod(seconds, 3600)
    minutes, seconds = divmod(remainder, 60)
    if hours:
        return f"{hours:02d}:{minutes:02d}:{seconds:02d}"
    return f"{minutes:02d}:{seconds:02d}"


def _text_height(values: list[Any], *, minimum: int = 32) -> int:
    lines = 1
    for raw in values:
        text = str(raw or "")
        explicit = text.count("\n") + 1
        wrapped = max(1, math.ceil(max((len(part) for part in text.split("\n")), default=0) / 20))
        lines = max(lines, explicit, wrapped)
    return min(240, max(minimum, 18 * lines + 10))


def _dimension_text(value: dict[str, Any]) -> str:
    text = value["fact"].strip()
    return "" if text == "未观察到" else text


def build_sheet_layout(analysis: dict[str, Any]) -> dict[str, Any]:
    """Purely derive all values, ranges and dimensions from validated analysis."""
    validate_analysis(analysis)
    slices = analysis["slices"]
    column_count = len(slices) + 1
    last_column = column_letter(column_count)

    matrix: list[list[Any]] = []
    matrix.append([TITLE] + [""] * len(slices))
    matrix.append(
        ["时间轴"]
        + [
            f"{_format_seconds(item['start'])}–{_format_seconds(item['end'])}"
            for item in slices
        ]
    )

    rows: dict[str, int] = {}
    for dimension in DIMENSION_KEYS:
        rows[dimension] = len(matrix) + 1
        values = []
        for item in slices:
            value = _dimension_text(item["dimensions"][dimension])
            if dimension == "阶段目标":
                value = "\n".join(
                    part for part in (item["stage_range"]["name"], value) if part
                )
            values.append(value)
        matrix.append([DIMENSION_LABELS[dimension]] + values)

    merges = [f"A1:{last_column}1"]
    stage_row = rows["阶段目标"]
    group_start = 0
    while group_start < len(slices):
        group_end = group_start
        stage_id = slices[group_start]["stage_range"]["stage_id"]
        while (
            group_end + 1 < len(slices)
            and slices[group_end + 1]["stage_range"]["stage_id"] == stage_id
        ):
            group_end += 1
        if group_end > group_start:
            merges.append(
                f"{column_letter(group_start + 2)}{stage_row}:"
                f"{column_letter(group_end + 2)}{stage_row}"
            )
            matrix[stage_row - 1][group_start + 1] = (
                "\n".join(
                    part
                    for part in (
                        slices[group_start]["stage_range"]["name"],
                        _dimension_text(
                            slices[group_start]["dimensions"]["阶段目标"]
                        ),
                    )
                    if part
                )
            )
            for position in range(group_start + 1, group_end + 1):
                matrix[stage_row - 1][position + 1] = ""
        group_start = group_end + 1

    full_range = f"A1:{last_column}{len(matrix)}"
    styles = [
        {
            "range": f"A1:{last_column}1",
            "style": {
                "backColor": MAIN_YELLOW,
                "foreColor": TITLE_FORE,
                "font": {"bold": True, "fontSize": "36pt/1.5"},
                "hAlign": 1,
                "vAlign": 1,
            },
        },
        {
            "range": f"A2:{last_column}2",
            "style": {
                "backColor": DARK_GREY,
                "foreColor": WHITE,
                "font": {"bold": True},
                "hAlign": 1,
                "vAlign": 1,
            },
        },
        {
            "range": f"A{stage_row}:{last_column}{stage_row}",
            "style": {"backColor": LIGHT_YELLOW, "font": {"bold": True}, "vAlign": 1},
        },
    ]
    for dimension, row in rows.items():
        styles.append(
            {
                "range": f"A{row}:A{row}",
                "style": {
                    "backColor": DARK_GREY,
                    "foreColor": WHITE,
                    "font": {"bold": True},
                    "hAlign": 1,
                    "vAlign": 1,
                },
            }
        )
        if dimension == "剧情轴":
            styles.append(
                {
                    "range": f"B{row}:{last_column}{row}",
                    "style": {"backColor": LIGHT_YELLOW, "vAlign": 0},
                }
            )

    row_heights = {1: 58, 2: 36}
    for dimension, row in rows.items():
        row_heights[row] = _text_height(matrix[row - 1], minimum=42 if dimension == "剧情轴" else 32)

    return {
        "matrix": matrix,
        "merges": merges,
        "styles": styles,
        "base_style": {
            "range": full_range,
            "style": {
                "font": {"fontSize": "11pt/1.5"},
                "hAlign": 0,
                "vAlign": 0,
                "borderType": "FULL_BORDER",
                "borderColor": "#D9D9D9",
            },
        },
        "column_widths": {1: 240, **{index: 220 for index in range(2, column_count + 1)}},
        "row_heights": row_heights,
        "rows": rows,
        "timeline_column_count": len(slices),
        "column_count": column_count,
        "validation_points": {
            "title_cell": "A1",
            "read_range": full_range,
            "expected_title": TITLE,
            "expected_merges": list(merges),
            "expected_title_color": MAIN_YELLOW,
            "key_rows": dict(rows),
        },
    }


class LarkRunner:
    """Small subprocess boundary used by production and replaced by tests."""

    def __init__(self, *, timeout: float = DEFAULT_TIMEOUT, executable: str | None = None):
        self.timeout = timeout
        self.executable = executable or ("lark-cli.cmd" if os.name == "nt" else "lark-cli")

    def run(
        self,
        operation: str,
        args: list[str],
        *,
        output_path: str | Path | None = None,
    ) -> dict[str, Any]:
        command = [self.executable, *args]
        try:
            completed = subprocess.run(
                command,
                capture_output=True,
                text=True,
                encoding="utf-8",
                errors="replace",
                timeout=self.timeout,
                check=False,
            )
        except (OSError, subprocess.TimeoutExpired) as exc:
            raise LarkCliError(f"{operation} 调用失败: {exc}") from exc
        if completed.returncode != 0:
            detail = completed.stderr.strip() or completed.stdout.strip() or "无错误详情"
            raise LarkCliError(f"{operation} 调用失败({completed.returncode}): {detail}")
        try:
            payload = json.loads(completed.stdout)
        except json.JSONDecodeError as exc:
            raise LarkCliError(f"{operation} 未返回合法 JSON: {completed.stdout[:300]}") from exc
        if not isinstance(payload, dict):
            raise LarkCliError(f"{operation} 返回值必须是 JSON 对象")
        code = payload.get("code", 0)
        if code not in (0, None):
            raise LarkCliError(f"{operation} API 错误: {json.dumps(payload, ensure_ascii=False)}")
        return payload


class _CreationTrackingRunner:
    def __init__(self, delegate: Any):
        self.delegate = delegate
        self.created_sheet_id: str | None = None
        self.spreadsheet_token: str | None = None

    def run(
        self,
        operation: str,
        args: list[str],
        *,
        output_path: str | Path | None = None,
    ) -> dict[str, Any]:
        result = self.delegate.run(operation, args, output_path=output_path)
        if operation == "create":
            for value in args:
                match = re.search(r"/spreadsheets/([^/]+)/sheets_batch_update$", value)
                if match:
                    self.spreadsheet_token = match.group(1)
                    break
            self.created_sheet_id = _created_sheet_id(result)
            if not self.created_sheet_id:
                raise LarkCliError("addSheet 成功响应缺少 sheetId，无法保证失败清理")
        return result


def _spreadsheet_argument(value: str) -> list[str]:
    if "://" in value:
        return ["--url", value]
    return ["--spreadsheet-token", value]


def _spreadsheet_token(value: str, metadata: dict[str, Any] | None = None) -> str:
    if metadata:
        direct = metadata.get("spreadsheet_token")
        if direct:
            return str(direct)
        data = metadata.get("data")
        if isinstance(data, dict):
            spreadsheet = data.get("spreadsheet")
            if isinstance(spreadsheet, dict):
                nested = spreadsheet.get("spreadsheet", spreadsheet)
                if isinstance(nested, dict) and nested.get("token"):
                    return str(nested["token"])
    match = re.search(r"/sheets/([^/?#]+)", value)
    if match:
        return match.group(1)
    if re.fullmatch(r"[A-Za-z0-9_-]+", value):
        return value
    raise ValueError("无法从目标 URL/token 解析 spreadsheet token")


def _sheets(metadata: dict[str, Any]) -> list[dict[str, Any]]:
    candidates = metadata.get("sheets")
    if candidates is None and isinstance(metadata.get("data"), dict):
        candidates = metadata["data"].get("sheets")
    if isinstance(candidates, dict):
        candidates = candidates.get("sheets")
    return candidates if isinstance(candidates, list) else []


def _sheet_id(sheet: dict[str, Any]) -> str | None:
    return sheet.get("sheet_id") or sheet.get("sheetId")


def _sheet_title(sheet: dict[str, Any]) -> str | None:
    return sheet.get("title") or sheet.get("sheet_title")


def _sheet_grid_size(sheet: dict[str, Any]) -> tuple[int, int]:
    grid = sheet.get("grid_properties")
    if not isinstance(grid, dict):
        grid = {}
    rows = sheet.get("row_count") or grid.get("row_count") or DEFAULT_SHEET_ROWS
    columns = (
        sheet.get("column_count")
        or grid.get("column_count")
        or DEFAULT_SHEET_COLUMNS
    )
    return int(rows), int(columns)


def _created_sheet_id(payload: dict[str, Any]) -> str | None:
    data = payload.get("data")
    if not isinstance(data, dict):
        return None
    replies = data.get("replies")
    if not isinstance(replies, list):
        return None
    for reply in replies:
        if not isinstance(reply, dict):
            continue
        properties = (reply.get("addSheet") or {}).get("properties") or {}
        sheet_id = properties.get("sheetId") or properties.get("sheet_id")
        if sheet_id:
            return str(sheet_id)
    return None


def _dimension_insert_requests(
    sheet_id: str,
    row_count: int,
    column_count: int,
    *,
    current_rows: int = DEFAULT_SHEET_ROWS,
    current_columns: int = DEFAULT_SHEET_COLUMNS,
) -> list[dict[str, Any]]:
    requests: list[dict[str, Any]] = []

    def append_requests(major: str, current: int, target: int) -> None:
        remaining = target - current
        while remaining > 0:
            count = min(remaining, current, 5000)
            requests.append(
                {
                    "dimension": {
                        "sheetId": sheet_id,
                        "majorDimension": major,
                        "startIndex": current - count,
                        "endIndex": current,
                    },
                    "inheritStyle": "BEFORE",
                }
            )
            current += count
            remaining -= count

    if column_count > current_columns:
        append_requests("COLUMNS", current_columns, column_count)
    if row_count > current_rows:
        append_requests("ROWS", current_rows, row_count)
    return requests


def _delete_sheet(
    runner: Any,
    token: str,
    sheet_id: str,
    identity: str,
) -> None:
    runner.run(
        "delete",
        _api_args(
            "POST",
            f"/open-apis/sheets/v2/spreadsheets/{token}/sheets_batch_update",
            identity,
            {"requests": [{"deleteSheet": {"sheetId": sheet_id}}]},
        ),
    )


def _json_arg(payload: Any) -> str:
    text = json.dumps(payload, ensure_ascii=False, separators=(",", ":"))
    if len(text) > MAX_JSON_ARGUMENT_CHARS:
        raise ValueError(
            f"单次 JSON 参数长度 {len(text)} 超过明确上限 {MAX_JSON_ARGUMENT_CHARS}；"
            "请减小 --batch-size 或缩短单个单元格"
        )
    return text


def _api_args(method: str, path: str, identity: str, payload: dict[str, Any]) -> list[str]:
    return ["api", method, path, "--as", identity, "--data", _json_arg(payload)]


def _metadata_args(spreadsheet: str, identity: str) -> list[str]:
    return ["sheets", "+info", "--as", identity, *_spreadsheet_argument(spreadsheet)]


def _extract_merges(sheet: dict[str, Any]) -> set[str]:
    raw = sheet.get("merges") or sheet.get("merged_cells") or []
    result = set()
    for item in raw:
        if isinstance(item, str):
            result.add(item.split("!", 1)[-1])
        elif isinstance(item, dict):
            value = item.get("range") or item.get("merged_range")
            if value:
                result.add(str(value).split("!", 1)[-1])
                continue
            indexes = (
                item.get("start_row_index"),
                item.get("end_row_index"),
                item.get("start_column_index"),
                item.get("end_column_index"),
            )
            if all(isinstance(index, int) and index >= 0 for index in indexes):
                start_row, end_row, start_column, end_column = indexes
                result.add(
                    f"{column_letter(start_column + 1)}{start_row + 1}:"
                    f"{column_letter(end_column + 1)}{end_row + 1}"
                )
    return result


def _read_values(payload: dict[str, Any]) -> list[list[Any]]:
    values = payload.get("values")
    if isinstance(values, list):
        return values
    data = payload.get("data")
    if not isinstance(data, dict):
        return []
    values = data.get("values")
    if isinstance(values, list):
        return values
    value_range = data.get("valueRange") or data.get("value_range")
    if isinstance(value_range, dict) and isinstance(value_range.get("values"), list):
        return value_range["values"]
    return []


def _build_plan(layout: dict[str, Any], batch_size: int) -> dict[str, Any]:
    operations: list[dict[str, Any]] = [
        {"operation": "metadata", "purpose": "检查同名页签冲突"},
        {"operation": "create", "title": "<new-sheet-title>"},
    ]
    if (
        layout["column_count"] > DEFAULT_SHEET_COLUMNS
        or len(layout["matrix"]) > DEFAULT_SHEET_ROWS
    ):
        operations.append(
            {
                "operation": "expand",
                "rows": len(layout["matrix"]),
                "columns": layout["column_count"],
            }
        )
    matrix = layout["matrix"]
    for start in range(0, len(matrix), batch_size):
        operations.append(
            {
                "operation": "write",
                "rows": [start + 1, min(start + batch_size, len(matrix))],
                "values": matrix[start : start + batch_size],
            }
        )
    operations.extend({"operation": "merge", "range": value} for value in layout["merges"])
    operations.extend(
        {"operation": "column_width", "column": index, "pixels": width}
        for index, width in layout["column_widths"].items()
    )
    operations.append({"operation": "base_style", **layout["base_style"]})
    operations.extend({"operation": "local_style", **style} for style in layout["styles"])
    operations.extend(
        {"operation": "row_height", "row": row, "pixels": height}
        for row, height in layout["row_heights"].items()
    )
    operations.extend(
        [
            {"operation": "read", "range": layout["validation_points"]["read_range"]},
            {"operation": "metadata", "purpose": "回读合并信息"},
            {"operation": "export", "extension": "xlsx", "temporary": True},
        ]
    )
    return {
        "dry_run": True,
        "limits": {
            "batch_size": batch_size,
            "max_batch_size": MAX_BATCH_SIZE,
            "max_json_argument_chars": MAX_JSON_ARGUMENT_CHARS,
        },
        "layout": layout,
        "operation_order": [
            "metadata",
            "create",
            "expand",
            "write",
            "merge",
            "column_width",
            "base_style",
            "local_style",
            "row_height",
            "read",
            "metadata",
            "export",
        ],
        "operations": operations,
    }


def verify_export_xlsx(
    path: str | Path,
    sheet_title: str,
    layout: dict[str, Any],
) -> dict[str, Any]:
    """Verify all required exported visual properties."""
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise VerificationError(
            "缺少 openpyxl，无法完成正式写回所需的 xlsx 样式验证"
        ) from exc

    workbook = load_workbook(path, read_only=False, data_only=True)
    try:
        if sheet_title not in workbook.sheetnames:
            raise VerificationError(f"导出文件缺少页签: {sheet_title}")
        worksheet = workbook[sheet_title]
        actual_merges = {str(value) for value in worksheet.merged_cells.ranges}
        missing = set(layout["merges"]) - actual_merges
        if missing:
            raise VerificationError(f"导出文件关键合并缺失: {sorted(missing)}")
        long_text_rows = {
            row
            for dimension, row in layout["rows"].items()
            if layout["row_heights"][row]
            > (42 if dimension == "剧情轴" else 32)
        }
        for row, pixels in layout["row_heights"].items():
            actual = worksheet.row_dimensions[row].height
            expected_points = pixels * (2 / 3)
            if actual is None or abs(float(actual) - expected_points) > 2:
                label = "长文本行高" if row in long_text_rows else "行高"
                raise VerificationError(
                    f"导出文件{label}错误（第 {row} 行）: "
                    f"{actual!r}，期望约 {expected_points}"
                )
        title = worksheet["A1"]
        fill_color = getattr(title.fill.fgColor, "rgb", None)
        if not isinstance(fill_color, str) or fill_color[-6:].upper() != MAIN_YELLOW[1:]:
            raise VerificationError(f"导出文件标题底色错误: {fill_color!r}")
        font_color = getattr(title.font.color, "rgb", None)
        if not isinstance(font_color, str) or font_color[-6:].upper() != TITLE_FORE[1:]:
            raise VerificationError(f"导出文件标题字色错误: {font_color!r}")
        if title.font.sz is None or abs(float(title.font.sz) - 36.0) > 0.1:
            raise VerificationError(f"导出文件标题字号错误: {title.font.sz!r}")
        for column, pixels in layout["column_widths"].items():
            letter = column_letter(column)
            actual_width = worksheet.column_dimensions[letter].width
            expected_width = excel_width_for_pixels(pixels)
            if actual_width is None or abs(float(actual_width) - expected_width) > 1.5:
                raise VerificationError(
                    f"导出文件时间轴与维度列宽错误（{letter} 列）: "
                    f"{actual_width!r}，期望约 {expected_width:.2f}"
                )
        return {
            "available": True,
            "passed": True,
            "checks": [
                "key_merges",
                "title_timeline_dimension_row_heights",
                "long_text_row_heights",
                "title_fill",
                "title_font_color",
                "title_font_size",
                "timeline_and_dimension_column_widths",
            ],
        }
    finally:
        workbook.close()


def _write_analysis_unprotected(
    analysis: dict[str, Any],
    spreadsheet: str,
    sheet_title: str,
    *,
    identity: str = "user",
    runner: Any | None = None,
    dry_run: bool = False,
    preflight: bool = False,
    batch_size: int = DEFAULT_BATCH_SIZE,
) -> dict[str, Any]:
    """Validate, create a new tab, write it, and verify the exported result."""
    validate_analysis(analysis)
    if not isinstance(sheet_title, str) or not sheet_title.strip():
        raise ValueError("new sheet title 必须是非空文本")
    if not 1 <= batch_size <= MAX_BATCH_SIZE:
        raise ValueError(f"batch size 必须在 1 到 {MAX_BATCH_SIZE} 之间")
    if dry_run and preflight:
        raise ValueError("--preflight 与 --dry-run 不能同时使用")
    layout = build_sheet_layout(analysis)
    plan = _build_plan(layout, batch_size)
    plan["sheet_title"] = sheet_title
    plan["spreadsheet"] = spreadsheet
    plan["identity"] = identity
    if dry_run:
        return plan

    runner = runner or LarkRunner()
    metadata = runner.run("metadata", _metadata_args(spreadsheet, identity))
    if any(_sheet_title(sheet) == sheet_title for sheet in _sheets(metadata)):
        raise SheetTitleConflict(
            f"目标表已存在同名页签“{sheet_title}”；只支持新建页签，"
            "请提供新的唯一页签名或取消"
        )
    if preflight:
        return {
            "preflight": True,
            "title_available": True,
            "sheet_title": sheet_title,
            "spreadsheet_token": _spreadsheet_token(spreadsheet, metadata),
            "identity": identity,
            "operations": ["metadata"],
        }
    token = _spreadsheet_token(spreadsheet, metadata)
    create_payload = {
        "requests": [
            {
                "addSheet": {
                    "properties": {
                        "title": sheet_title,
                    }
                }
            }
        ]
    }
    create_result = runner.run(
        "create",
        _api_args(
            "POST",
            f"/open-apis/sheets/v2/spreadsheets/{token}/sheets_batch_update",
            identity,
            create_payload,
        ),
    )
    sheet_id = _created_sheet_id(create_result)
    if not sheet_id:
        raise LarkCliError("addSheet 成功响应缺少 sheetId")
    created_metadata = runner.run("metadata", _metadata_args(spreadsheet, identity))
    matching = [sheet for sheet in _sheets(created_metadata) if _sheet_title(sheet) == sheet_title]
    if (
        len(matching) != 1
        or not _sheet_id(matching[0])
        or str(_sheet_id(matching[0])) != sheet_id
    ):
        raise LarkCliError("创建后 metadata 未找到唯一的新页签")
    created_sheet = matching[0]
    current_rows, current_columns = _sheet_grid_size(created_sheet)
    insert_requests = _dimension_insert_requests(
        sheet_id,
        len(layout["matrix"]),
        layout["column_count"],
        current_rows=current_rows,
        current_columns=current_columns,
    )
    for insert_payload in insert_requests:
        runner.run(
            "expand",
            _api_args(
                "POST",
                f"/open-apis/sheets/v2/spreadsheets/{token}/insert_dimension_range",
                identity,
                insert_payload,
            ),
        )
    last_column = column_letter(layout["column_count"])

    matrix = layout["matrix"]
    for start in range(0, len(matrix), batch_size):
        values = matrix[start : start + batch_size]
        end = start + len(values)
        cell_range = f"{sheet_id}!A{start + 1}:{last_column}{end}"
        runner.run(
            "write",
            [
                "sheets",
                "+write",
                "--as",
                identity,
                "--spreadsheet-token",
                token,
                "--range",
                cell_range,
                "--values",
                _json_arg(values),
            ],
        )

    for cell_range in layout["merges"]:
        runner.run(
            "merge",
            _api_args(
                "POST",
                f"/open-apis/sheets/v2/spreadsheets/{token}/merge_cells",
                identity,
                {"range": f"{sheet_id}!{cell_range}", "mergeType": "MERGE_ALL"},
            ),
        )

    width_groups = [(1, 1, layout["column_widths"][1])]
    if layout["column_count"] > 1:
        width_groups.append((2, layout["column_count"], layout["column_widths"][2]))
    for start, end, pixels in width_groups:
        runner.run(
            "column_width",
            _api_args(
                "PUT",
                f"/open-apis/sheets/v2/spreadsheets/{token}/dimension_range",
                identity,
                {
                    "dimension": {
                        "sheetId": sheet_id,
                        "majorDimension": "COLUMNS",
                        "startIndex": start,
                        "endIndex": end,
                    },
                    "dimensionProperties": {"fixedSize": pixels, "visible": True},
                },
            ),
        )

    base = layout["base_style"]
    runner.run(
        "base_style",
        _api_args(
            "PUT",
            f"/open-apis/sheets/v2/spreadsheets/{token}/styles_batch_update",
            identity,
            {"data": [{"ranges": [f"{sheet_id}!{base['range']}"], "style": base["style"]}]},
        ),
    )
    for style in layout["styles"]:
        runner.run(
            "local_style",
            _api_args(
                "PUT",
                f"/open-apis/sheets/v2/spreadsheets/{token}/styles_batch_update",
                identity,
                {"data": [{"ranges": [f"{sheet_id}!{style['range']}"], "style": style["style"]}]},
            ),
        )

    for row, pixels in layout["row_heights"].items():
        runner.run(
            "row_height",
            _api_args(
                "PUT",
                f"/open-apis/sheets/v2/spreadsheets/{token}/dimension_range",
                identity,
                {
                    "dimension": {
                        "sheetId": sheet_id,
                        "majorDimension": "ROWS",
                        "startIndex": row,
                        "endIndex": row,
                    },
                    "dimensionProperties": {"fixedSize": pixels, "visible": True},
                },
            ),
        )

    read_result = runner.run(
        "read",
        [
            "sheets",
            "+read",
            "--as",
            identity,
            "--spreadsheet-token",
            token,
            "--range",
            f"{sheet_id}!{layout['validation_points']['read_range']}",
        ],
    )
    values = _read_values(read_result)
    if not values or not values[0] or values[0][0] != TITLE:
        raise VerificationError("写后回读未找到正确总标题")
    for dimension, row in layout["rows"].items():
        if (
            len(values) < row
            or not values[row - 1]
            or values[row - 1][0] != DIMENSION_LABELS[dimension]
        ):
            raise VerificationError(f"写后回读缺少关键维度行: {dimension}")

    final_metadata = runner.run("metadata", _metadata_args(spreadsheet, identity))
    final_matches = [sheet for sheet in _sheets(final_metadata) if _sheet_id(sheet) == sheet_id]
    if len(final_matches) != 1:
        raise VerificationError("写后 metadata 未找到新页签")
    actual_merges = _extract_merges(final_matches[0])
    expected_merges = set(layout["merges"])
    if not expected_merges.issubset(actual_merges):
        raise VerificationError(
            f"写后 metadata 合并校验失败，缺少: {sorted(expected_merges - actual_merges)}"
        )

    export_path: Path | None = None
    try:
        export_root = Path.cwd()
        with tempfile.NamedTemporaryFile(
            suffix=".xlsx", delete=False, dir=export_root
        ) as temporary:
            export_path = Path(temporary.name)
        export_path.unlink()
        cli_export_path = export_path.relative_to(export_root).as_posix()
        runner.run(
            "export",
            [
                "sheets",
                "+export",
                "--as",
                identity,
                "--spreadsheet-token",
                token,
                "--file-extension",
                "xlsx",
                "--output-path",
                cli_export_path,
            ],
            output_path=export_path,
        )
        if not export_path.is_file() or export_path.stat().st_size == 0:
            raise VerificationError("导出命令未生成非空 xlsx 文件")
        try:
            export_verification = verify_export_xlsx(export_path, sheet_title, layout)
        except VerificationError:
            raise
        except Exception as exc:
            raise VerificationError(f"xlsx 样式验证失败: {exc}") from exc
        if not export_verification.get("passed"):
            raise VerificationError("xlsx 样式验证失败")
    finally:
        if export_path is not None:
            try:
                export_path.unlink()
            except FileNotFoundError:
                pass

    return {
        "spreadsheet_token": token,
        "sheet_id": sheet_id,
        "sheet_title": sheet_title,
        "layout": {
            "rows": len(layout["matrix"]),
            "columns": layout["column_count"],
            "timeline_columns": layout["timeline_column_count"],
        },
        "verification": {
            "passed": True,
            "content_readback": True,
            "metadata_merges": sorted(actual_merges),
            "xlsx": export_verification,
        },
    }


def write_analysis(
    analysis: dict[str, Any],
    spreadsheet: str,
    sheet_title: str,
    *,
    identity: str = "user",
    runner: Any | None = None,
    dry_run: bool = False,
    preflight: bool = False,
    batch_size: int = DEFAULT_BATCH_SIZE,
) -> dict[str, Any]:
    """Run the write transaction and delete this invocation's tab on failure."""
    delegate = runner or LarkRunner()
    tracking_runner = _CreationTrackingRunner(delegate)
    try:
        return _write_analysis_unprotected(
            analysis,
            spreadsheet,
            sheet_title,
            identity=identity,
            runner=tracking_runner,
            dry_run=dry_run,
            preflight=preflight,
            batch_size=batch_size,
        )
    except Exception as cause:
        if not tracking_runner.created_sheet_id:
            raise
        cleanup_error = None
        try:
            token = tracking_runner.spreadsheet_token or _spreadsheet_token(spreadsheet)
            _delete_sheet(
                delegate,
                token,
                tracking_runner.created_sheet_id,
                identity,
            )
        except Exception as exc:
            cleanup_error = exc
        raise WriteTransactionError(cause, cleanup_error) from cause


def main(argv: list[str] | None = None, *, runner: Any | None = None) -> int:
    parser = argparse.ArgumentParser(description="将游戏前期体验拆解写入新的飞书页签")
    parser.add_argument("analysis", help="已完成的 analysis.json")
    parser.add_argument("spreadsheet", help="目标 spreadsheet URL 或 token")
    parser.add_argument("sheet_title", help="新页签标题；同名时拒绝写入")
    parser.add_argument("--identity", choices=("user", "bot"), default="user")
    mode = parser.add_mutually_exclusive_group()
    mode.add_argument(
        "--preflight",
        action="store_true",
        help="只读调用 sheets +info/metadata，确认新页签名唯一",
    )
    mode.add_argument("--dry-run", action="store_true", help="仅输出完整操作计划，不调用网络")
    parser.add_argument("--batch-size", type=int, default=DEFAULT_BATCH_SIZE)
    parser.add_argument("--timeout", type=float, default=DEFAULT_TIMEOUT)
    args = parser.parse_args(argv)
    try:
        source = Path(args.analysis)
        analysis = json.loads(source.read_text(encoding="utf-8"))
        active_runner = runner or LarkRunner(timeout=args.timeout)
        result = write_analysis(
            analysis,
            args.spreadsheet,
            args.sheet_title,
            identity=args.identity,
            runner=active_runner,
            dry_run=args.dry_run,
            preflight=args.preflight,
            batch_size=args.batch_size,
        )
        sys.stdout.write(json.dumps(result, ensure_ascii=False, indent=2) + "\n")
        return 0
    except WriteTransactionError as exc:
        cleanup = (
            {"status": "failed", "message": str(exc.cleanup_error)}
            if exc.cleanup_error is not None
            else {"status": "success", "message": "已删除本次新建页签"}
        )
        payload = {
            "error": {
                "type": type(exc).__name__,
                "message": str(exc),
                "cause": {
                    "type": type(exc.cause).__name__,
                    "message": str(exc.cause),
                },
                "cleanup": cleanup,
            }
        }
        sys.stderr.write(json.dumps(payload, ensure_ascii=False, indent=2) + "\n")
        return 2
    except (
        OSError,
        UnicodeError,
        json.JSONDecodeError,
        ValueError,
        LarkCliError,
        SheetTitleConflict,
        VerificationError,
    ) as exc:
        emit_json_error(exc)
        return 2


if __name__ == "__main__":
    raise SystemExit(main())
